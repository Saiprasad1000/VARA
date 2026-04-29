from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Sum, Count
from django.core.paginator import Paginator
from django.utils import timezone
from datetime import timedelta, datetime
from decimal import Decimal
from .common_importers import admin_required
from User.models import Order

import io


# ── Helpers ───────────────────────────────────────────────────────────────────

def _apply_date_filter(qs, period, date_from, date_to):
    now = timezone.now()
    if period == 'daily':
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        qs = qs.filter(created_at__gte=start)
    elif period == 'weekly':
        start = now - timedelta(days=now.weekday())
        start = start.replace(hour=0, minute=0, second=0, microsecond=0)
        qs = qs.filter(created_at__gte=start)
    elif period == 'monthly':
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        qs = qs.filter(created_at__gte=start)
    elif period == 'yearly':
        start = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        qs = qs.filter(created_at__gte=start)
    elif period == 'custom' and date_from and date_to:
        try:
            df = datetime.strptime(date_from, '%Y-%m-%d')
            dt = datetime.strptime(date_to, '%Y-%m-%d')
            qs = qs.filter(created_at__date__gte=df.date(), created_at__date__lte=dt.date())
        except ValueError:
            pass
    return qs


def _get_orders(request):
    period = request.GET.get('period', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    qs = Order.objects.select_related('user', 'coupon').order_by('-created_at')
    qs = _apply_date_filter(qs, period, date_from, date_to)
    return qs, period, date_from, date_to


# ── Main view ─────────────────────────────────────────────────────────────────

@admin_required
def sales_report(request):
    orders, period, date_from, date_to = _get_orders(request)

    # Summary metrics
    total_orders = orders.count()
    valid_orders = orders.exclude(status__in=['Cancelled', 'Returned'])
    totals = valid_orders.aggregate(
        total_sales=Sum('total_amount'),
        total_discount=Sum('discount_amount'),
    )
    total_sales = totals['total_sales'] or Decimal('0.00')
    total_discount = totals['total_discount'] or Decimal('0.00')
    net_revenue = total_sales  # total_amount already stores post-discount paid amount

    # Paginate
    paginator = Paginator(orders, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'orders': page_obj,
        'period': period,
        'date_from': date_from,
        'date_to': date_to,
        'total_orders': total_orders,
        'total_sales': total_sales,
        'total_discount': total_discount,
        'net_revenue': net_revenue,
    }
    return render(request, 'sales_report.html', context)


# ── PDF Download ──────────────────────────────────────────────────────────────

@admin_required
def download_sales_pdf(request):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    orders, period, date_from, date_to = _get_orders(request)

    total_orders = orders.count()
    valid_orders = orders.exclude(status__in=['Cancelled', 'Returned'])
    totals = valid_orders.aggregate(
        total_sales=Sum('total_amount'),
        total_discount=Sum('discount_amount'),
    )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=1*cm, leftMargin=1*cm,
                            topMargin=1.5*cm, bottomMargin=1*cm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('title', parent=styles['Title'], fontSize=18,
                                  textColor=colors.HexColor('#6B3E26'), spaceAfter=6)
    sub_style = ParagraphStyle('sub', parent=styles['Normal'], fontSize=10,
                                textColor=colors.grey, spaceAfter=12)

    elements = []

    # Title
    elements.append(Paragraph('VARA – Sales Report', title_style))
    period_label = period.capitalize() if period else 'All Time'
    if period == 'custom' and date_from and date_to:
        period_label = f'{date_from} to {date_to}'
    elements.append(Paragraph(f'Period: {period_label}', sub_style))
    elements.append(Spacer(1, 0.3*cm))

    # Summary
    summary_data = [
        ['Total Orders', 'Total Sales (₹)', 'Total Discount (₹)', 'Net Revenue (₹)'],
        [
            str(total_orders),
            f"{totals['total_sales'] or 0:.2f}",
            f"{totals['total_discount'] or 0:.2f}",
            f"{totals['total_sales'] or 0:.2f}",
        ]
    ]
    summary_table = Table(summary_data, colWidths=[4*cm, 5*cm, 5*cm, 5*cm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F7F4ED')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#6B3E26')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E7E2D9')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FBF9F4')]),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.5*cm))

    # Orders table
    headers = ['Order ID', 'Date', 'Customer', 'Payment', 'Discount (₹)', 'Total (₹)', 'Status']
    rows = [headers]
    for order in orders[:500]:  # cap at 500 rows in PDF
        rows.append([
            f'#{order.id}',
            order.created_at.strftime('%b %d, %Y'),
            f'{order.user.first_name} {order.user.last_name}',
            order.payment_method,
            f'{order.discount_amount:.2f}',
            f'{order.total_amount:.2f}',
            order.status,
        ])

    col_widths = [2.5*cm, 3*cm, 5*cm, 3*cm, 3.5*cm, 3.5*cm, 3.5*cm]
    t = Table(rows, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6B3E26')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#E7E2D9')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FBF9F4')]),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(t)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="vara_sales_report.pdf"'
    return response


# ── Excel Download ────────────────────────────────────────────────────────────

@admin_required
def download_sales_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    orders, period, date_from, date_to = _get_orders(request)
    total_orders = orders.count()
    valid_orders = orders.exclude(status__in=['Cancelled', 'Returned'])
    totals = valid_orders.aggregate(
        total_sales=Sum('total_amount'),
        total_discount=Sum('discount_amount'),
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sales Report'

    # ── Colour palette ───────
    brown = 'FF6B3E26'
    cream = 'FFF7F4ED'
    light = 'FFFBF9F4'
    white = 'FFFFFFFF'

    header_font = Font(bold=True, color='FFFFFFFF', size=11)
    brown_fill = PatternFill('solid', fgColor=brown)
    cream_fill = PatternFill('solid', fgColor=cream)
    light_fill = PatternFill('solid', fgColor=light)
    center = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='FFE7E2D9')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Title ─────────────────
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = 'VARA – Sales Report'
    title_cell.font = Font(bold=True, size=16, color=brown[2:])
    title_cell.alignment = center
    ws.row_dimensions[1].height = 28

    period_label = period.capitalize() if period else 'All Time'
    if period == 'custom' and date_from and date_to:
        period_label = f'{date_from} to {date_to}'
    ws.merge_cells('A2:G2')
    ws['A2'].value = f'Period: {period_label}'
    ws['A2'].alignment = center
    ws.row_dimensions[2].height = 18

    # ── Summary ───────────────
    ws.row_dimensions[4].height = 20
    summary_headers = ['Total Orders', 'Total Sales', 'Total Discount', 'Net Revenue']
    summary_values = [
        total_orders,
        float(totals['total_sales'] or 0),
        float(totals['total_discount'] or 0),
        float(totals['total_sales'] or 0),
    ]
    for col, (h, v) in enumerate(zip(summary_headers, summary_values), start=1):
        hc = ws.cell(row=4, column=col, value=h)
        hc.font = Font(bold=True, size=10)
        hc.fill = cream_fill
        hc.alignment = center
        hc.border = border

        vc = ws.cell(row=5, column=col, value=v)
        vc.font = Font(size=10)
        vc.alignment = center
        vc.border = border
        if col > 1:
            vc.number_format = '₹#,##0.00'

    # ── Orders table ──────────
    headers = ['Order ID', 'Date', 'Customer', 'Email', 'Payment', 'Discount (₹)', 'Total (₹)', 'Status']
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=7, column=col, value=h)
        cell.font = header_font
        cell.fill = brown_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[7].height = 20

    for r, order in enumerate(orders, start=8):
        row_fill = light_fill if r % 2 == 0 else PatternFill('solid', fgColor=white)
        data = [
            f'#{order.id}',
            order.created_at.strftime('%b %d, %Y'),
            f'{order.user.first_name} {order.user.last_name}',
            order.user.email,
            order.payment_method,
            float(order.discount_amount),
            float(order.total_amount),
            order.status,
        ]
        for col, val in enumerate(data, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.fill = row_fill
            cell.alignment = center
            cell.border = border
            if col in (6, 7):
                cell.number_format = '₹#,##0.00'

    # Column widths
    col_widths = [12, 14, 20, 28, 12, 16, 14, 14]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="vara_sales_report.xlsx"'
    return response
